import pandas as pd
import numpy as np
from datetime import datetime
import scipy.stats as st #計算統計和信賴區間
import pymysql #連線sql
from openpyxl import load_workbook #寫入excel

def compute_deltas(time_list, value_list, expected_interval=5, method="rescale"):
    """
    time_list: datetime list
    value_list: 對應剩餘車位數
    expected_interval: 預期間隔分鐘數，預設 5
    method: "rescale", "drop", "interpolate"
    """
    df = pd.DataFrame({"time": time_list, "value": value_list})
    df = df.sort_values("time").reset_index(drop=True)
    
    # 計算相鄰時間間隔（分鐘）
    df["delta_minutes"] = df["time"].diff().dt.total_seconds() / 60.0
    
    # 計算原始變化量
    df["delta_raw"] = df["value"].diff()
    
    if method == "rescale":
        # 平攤成每 expected_interval 分鐘的變化量
        df["delta"] = df.apply(
            lambda row: (row["delta_raw"] / (row["delta_minutes"]/expected_interval))
            if pd.notnull(row["delta_raw"]) else None,
            axis=1
        )
    elif method == "drop":
        # 只保留間隔接近預期的資料
        df = df[(df["delta_minutes"].isna()) | (df["delta_minutes"] <= expected_interval*1.5)]
        df["delta"] = df["delta_raw"]
    elif method == "interpolate":
        # 先補全每 expected_interval 的時間點
        full_time = pd.date_range(start=df["time"].iloc[0], end=df["time"].iloc[-1], freq=f"{expected_interval}T")
        df_full = pd.DataFrame({"time": full_time})
        df_full = df_full.merge(df[["time", "value"]], on="time", how="left")
        df_full["value"] = df_full["value"].interpolate(method="linear")
        df_full["delta"] = df_full["value"].diff()
        df = df_full
    else:
        raise ValueError("method 必須是 'rescale', 'drop', 或 'interpolate'")
    
    return df[["time", "value", "delta"]]

# -----------------------
# SQL資料抓取(密碼自行修正)
connect = pymysql.connect(
    host='localhost',
    user='root',
    password='yourpassword',
    db='parking_info',
    charset='utf8mb4',
    cursorclass=pymysql.cursors.DictCursor
)

#指定停車場編號、星期和時間範圍
llist=['TPE0959','TPE0099','TPE0006','TPE0096','TPE0742',
       'TPE0824','TPE0005','TPE0160','TPE0181','TPE0377',
       'TPE0680']
try:
    for no in llist:
        for wk in range(1,8):
            for x in range(6,22):
                try:
                    with connect.cursor() as cursor:
                        strtime=f"{x:02d}:00:00"
                        endtime=f"{x+1:02d}:04:00"
                        sql = f'''
                        SELECT `站名`,`剩餘車位`,`更新時間` FROM parking_lot
                        WHERE DAYOFWEEK(`更新時間`) = {wk} 
                        AND id = "{no}"
                        AND TIME(`更新時間`) >= "{strtime}"
                        AND TIME(`更新時間`) <= "{endtime}"
                        '''
                        cursor.execute(sql)
                        ex=cursor.fetchall()
                        
                        count=[]
                        for i in ex:
                            last=i["剩餘車位"]
                            time=i["更新時間"]
                            name=i["站名"]
                            count.extend([last,time])
                            
                        #print(count)
                        #print(len(count))
                        #print(name)
                        timelist=count[1::2]
                        spacelist=count[0::2]
                        #print(timelist)
                        #print(spacelist)
                        #time=ex['更新時間']
                        #print(time)
                        
                        df_deltas = compute_deltas(timelist, spacelist, expected_interval=5, method="drop")
                        #print(df_deltas)
                        
                        # 計算t分配信賴區間
                        df_deltas_spc = df_deltas['delta']
                        clean = df_deltas_spc[~np.isnan(df_deltas_spc)]
                        
                        n = len(clean)
                        mean = clean.mean()
                        s = np.std(clean,ddof=1) #樣本標準差
                        dof = n-1
                        alpha = 0.05
                        t = st.t.ppf(1-alpha/2,dof)
                        print(f"id:{no}")
                        print(f"停車場:{name}")
                        print(f"星期:{wk-1}")
                        print(f"起始時間:{strtime}")
                        print(f"結束時間:{endtime}")
                        print(f"資料值：{n}")
                        print(f"樣本平均數:{mean}")
                        print(f"樣本標準差:{s}")
                        print(f"自由度:{dof}")
                        print(f"顯著水準:{alpha}")
                        print(f"t值:{t:.2f}")
                        
                        se = s/(np.sqrt(n))
                        high_ci = mean + t*se
                        low_ci = mean - t*se
                        print(f"信賴區間下限:{low_ci:.2f}")
                        print(f"信賴區間上限:{high_ci:.2f}")
                        
                        '''
                        帶入樣本數是每五分鐘的車流進出量，U輛 /5分
                        帶入為，U/5輛 /分
                        轉換為，1/(U/5)分 /輛，也就是(5/U分 /輛)
                        求得此信賴區間可以說等待一輛出車有95%信心會在5/L~5/U分之間
                        注意！若U或L小於0，代表車輛還在流入不建議等待；大於才是有效數字
                        '''
                        low_wait = f"{5/low_ci:.1f}"
                        up_wait = f"{5/high_ci:.1f}"
                        good=f'有95%信心預計車位釋出時間在於{up_wait}分至{low_wait}分'
                        normal=f"可能預計等待至少{up_wait}分"
                        bad="預計等待時間過長，建議前往...停車場"
                        
                        if low_ci > 0 and high_ci > 0 :
                            result=good
                            print(result)
                        elif high_ci > 0 :
                            result=normal
                            print(result)
                        else:
                            result=bad
                            print(result)
                        print()
                    
                        # 寫入excel
                        wb = load_workbook("D:/your_excel.xlsx")
                        ws = wb.active
                        '''
                        建立完一次欄位名就好
                        ws['A1']="停車場"
                        ws['B1']="名稱"
                        ws['C1']="星期"
                        ws['D1']="起始時間"
                        ws['E1']="結束時間"
                        ws['F1']="資料值"
                        ws['G1']="樣本平均數"
                        ws['H1']="樣本標準差"
                        ws['I1']="自由度"
                        ws['J1']="顯著水準"
                        ws['K1']="t值"
                        ws['L1']="信賴區間下限"
                        ws['M1']="信賴區間上限"
                        ws['N1']="較短等待時間"
                        ws['O1']="較長等待時間"
                        ws['P1']="建議結果"
                        '''
                        ws.append([no,name,wk-1,strtime,endtime,
                                   n,mean,s,dof,alpha,f'{t:.2f}',
                                   f'{low_ci:.3f}',f'{high_ci:.3f}',
                                   up_wait,low_wait,result])
                        
                        wb.save("D:/analysis project/data/parkinglot_CI.xlsx")
                except Exception as e:
                    print(f"處理週{wk-1} 時間{strtime}-{endtime}時發生錯誤: {e}")
                    pass
finally:
    connect.close()

